# ------------------------------- /usr/bin/g++-7 ------------------------------#
# ------------------------------- coding: utf-8 -------------------------------#
# Criado por:   Jean Marcelo Mira Junior
#               Victor Philos Donato Luiz da Silva
# Versão: 1.0
# Criado em: 22/05/2021
# Sistema operacional: Linux - Ubuntu 30.04.1 LTS
# Python 3
# ------------------------------ Pacotes --------------------------------------#
import os
import openpyxl
import matplotlib.pyplot as plt
import pickle

from pybrain3.tools.shortcuts import buildNetwork
from pybrain3.datasets import SupervisedDataSet
from pybrain3.supervised.trainers import BackpropTrainer
# -----------------------------------------------------------------------------#


def analise(nomeDoArquivo):
    interacoes = 2500  # Número de interações
    fileObject = open('redesalva.xml', 'wb')  # Abre arquivo

    # Definição da rede
    # rede = buildNetwork(Neurônios camada de entrada, Neurônios camada oculta, Neurônios camada de  saída, Unidade de Bias(valor unitário conectado a um neurônio))
    base = SupervisedDataSet(1, 1)
    rede = buildNetwork(base.indim, 120, 350, 120, 300,
                        120, 350, 500, 400, 100, base.outdim)

    # Manipulação de arquivos
    diretorio = os.getcwd()  # Salva os nome caminho do diretório

    arquivo = openpyxl.load_workbook(
        diretorio + "/" + nomeDoArquivo, data_only=True)
    folha = arquivo.active

    entradaTreino = []
    saidaTreino = []
    entradaTeste = []
    saidaTeste = []

    for i in range(1, folha.max_row):
        if((i % 5) == 0):
            entradaTreino.append(float(folha.cell(row=i+1, column=2).value))
            saidaTreino.append(float(folha.cell(row=i+1, column=3).value))
        else:
            entradaTeste.append(float(folha.cell(row=i+1, column=2).value))
            saidaTeste.append(float(folha.cell(row=i+1, column=3).value))

    normalEntradaTreino = []
    normalSaidaTreino = []
    normalEntradaTeste = []
    normalSaidaTeste = []

    # Normaliza os dados
    for i in range(len(entradaTreino)):
        normalEntradaTreino.append(
            (entradaTreino[i]-min(entradaTreino))/(max(entradaTreino)-min(entradaTreino)))
        normalSaidaTreino.append(
            (saidaTreino[i]-min(saidaTreino))/(max(saidaTreino)-min(saidaTreino)))

    for i in range(len(saidaTeste)):
        normalEntradaTeste.append(
            (entradaTeste[i]-min(entradaTeste))/(max(entradaTeste)-min(entradaTeste)))
        normalSaidaTeste.append(
            (saidaTeste[i]-min(saidaTeste))/(max(saidaTeste)-min(saidaTeste)))

    """ for i in range(len(entrada)):
        print(entrada[i], saida[i]) """

    # Alimentação da rede neural artificial

    for i in range(len(normalEntradaTreino)):
        base.addSample(([normalEntradaTreino[i]]), ([normalSaidaTreino[i]]))

    treinamento = BackpropTrainer(rede, dataset=base, verbose=True)

    erro = 0
    for i in range(1, interacoes):
        print("[", i, "]")
        erro += treinamento.train()
    erro /= interacoes

    pickle.dump(rede, fileObject)  # Guarda dados de treino
    fileObject.close()  # Fecha arquivo

    # Faz o gráfico dos dados de treino da saída real x saída da rede neural artificial
    dadoSaida = []
    for i in range(len(entradaTreino)):
        dadoSaida.append(rede.activate([normalEntradaTreino[i]]))

    plt.figure(figsize=(7, 5))
    plt.plot(dadoSaida, color='red', label='Saída treino rede neural')
    plt.plot(normalSaidaTreino, 'k--', label='Saída treino real')
    plt.grid(True)
    plt.title('Saída treino real x Saída treino rede neural')
    plt.text(1.5, 0.01, "Número de interações: 2500")
    plt.text(
        1.5, 0.06, "buildNetwork(1, 120, 350, 120, 300, 120, 350, 500, 400, 100, 1)")
    plt.text(1.5, 0.11, "Erro: " + str(erro))
    plt.legend()
    plt.savefig("Dados_treino_FINAL_2500.png")
    plt.show()

    # Avaliando a rede com os dados de teste, faz o gráfico dos dados de teste da saída real x saída da rede neural artificial
    """ dadoSaida = []
    for i in range(len(entradaTeste)):
        dadoSaida.append(rede.activate([normalEntradaTeste[i]]))

    plt.figure(figsize=(7, 5))
    plt.plot(normalSaidaTeste, color='green',
             label='Saída treino real')
    plt.plot(dadoSaida, color='red', label='Saída treino rede neural')
    plt.grid(True)
    plt.legend()
    plt.text(6.5, 0, "Erro: " + str(erro))
    plt.title('Saída teste real x Saída teste rede neural')
    plt.savefig("teste.png")
    plt.show()
 """


def aplicacao(min, max):
    fileObject = open('redesalva.xml', 'rb')
    rede = pickle.load(fileObject)

    # Interação com o usuário

    usuario = float(0)
    while(usuario != -1):
        usuario = float(input("Entrada: "))
        n = rede.activate(([float(usuario-min)/(max-min)]))
        print("Saida: ", (n[0] * (max-min) + min))


def retornaMinMax(nomeDoArquivo):
    # Manipulação de arquivos
    diretorio = os.getcwd()  # Salva os nome caminho do diretório

    arquivo = openpyxl.load_workbook(
        diretorio + "/" + nomeDoArquivo, data_only=True)
    folha = arquivo.active

    entradaTreino = []

    for i in range(1, folha.max_row):
        entradaTreino.append(float(folha.cell(row=i+1, column=2).value))

    return(min(entradaTreino), max(entradaTreino))
